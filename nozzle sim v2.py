
#huzel and huang
#Rocket Propulsion Elements
#www.grc.nasa.gov/
#http://www.aspirespace.org.uk
#Rao

import math
import sys
import openpyxl


def C_liquid(T):
    Tc=309.57
    Tr=T/Tc
    #isobaric specific heat capacity of the saturated liquid
    #eq 4.7 thermophysical properties
    #Aplicable from -90C to 30C
    b1=2.49973
    b2=0.023454
    b3=-3.80136
    b4=13.0945
    b5=-14.5180
    return(1000*b1*(1+b2*(1-Tr)**(-1)+b3*(1-Tr)+b4*(1-Tr)**2+b5*(1-Tr)**3))


def rho_liquid(T):
    Tc=309.57
    Tr=T/Tc
    #Density of saturated liquid
    #eq 4.2 thermophysical properties
    #Aplicable from -90C to 36C
    #constants
    rho_c=452.0 #critical density
    e=2.718281828459
    b1=1.72328
    b2=-.83950
    b3=.51060
    b4=-.10412
    return(rho_c*(e**(b1*(1-Tr)**(1/3)+b2*(1-Tr)**(2/3)+b3*(1-Tr)+b4*(1-Tr)**(4/3))))


def rho_vapor(T):
    Tc=309.57
    Tr=T/Tc
    #Density of saturated vapor
    #eq 4.3 thermophysical properties
    #Aplicable from -90C to 36C

    #constants
    rho_c=452.0 #critical density
    e=2.718281828459

    b1=-1.00900
    b2=-6.28792
    b3=7.50332
    b4=-7.90463
    b5=0.629427
    return(rho_c*(e**(b1*((1/Tr)-1)**(1/3)+b2*((1/Tr)-1)**(2/3)+b3*((1/Tr)-1)+b4*((1/Tr)-1)**(4/3)+b5*((1/Tr)-1)**(5/3))))

def DynViscosityLiquid(T):
    Tc = 309.57
    Tr = T / Tc
    b1=1.6089
    b2=2.0439
    b3=5.24
    b4=.0293423
    theta=(Tc-b3)/(T-b3)
    return b4*math.exp(b1*(theta-1)**(1/3)+b2*(theta-1)**(4/3))

def ConductivityLiquid(T):
    Tc = 309.57
    Tr = T / Tc
    b1=72.35
    b2=1.5
    b3=-3.5
    b4=4.5
    return b1*(1+b2*(1-Tr)**(1/3)+b3*(1-Tr)**(2/3)+b4*(1-Tr))

def Fahrenheit_to_Kelvin(F):
    return 273.5 + ((F - 32.0) * (5.0/9.0))


def Meters_to_Inches(inch):
    return 39.37007874*inch


defaultq=input("Accept defaults? ")
if defaultq=="yes" or defaultq=="":
    defaults=True
elif defaultq=="no":
    defaults=False
else:
    print("Invalid input")
    defaults=False

p1=6894.76*350 #p1 is chamber pressure
k = 1.236844 #property of your working fluid
T1=3057.93 #Fahrenheit_to_Kelvin(3539.93) #temperature in the chamber in kelvin
p3=6894.76*14.5 #free stream pressure outside nozzle -> Pa
p2=p3 #True for best performance
AverageMolecularWeight=0.02466018 #this is in kg. Not typical for molecular weight. Rs=Ru/MolarMass
LStar=1.67#meters

initNoxTemp=273

pi=3.14159265
R=8.314462 #gas constant
Rspecific=R/AverageMolecularWeight

pt=((2/(k+1))**(k/(k-1)))*p1 #57
Tt=(2*T1)/(k+1) #page 57
vt=math.sqrt(((2*k)/(k+1))*Rspecific*T1) #page 57


v2=math.sqrt(((2*k)/(k-1))*Rspecific*T1*(1-(p2/p1)**((k-1)/k))) #page 54
T2=T1*(p2/p1)**((k-1)/k)
M2=v2/math.sqrt(k*Rspecific*T2)

V1=Rspecific*T1/p1 #page 55
Vt=V1*((k+1)/2)**(1/(k-1)) #page 57
while True:
    print("""Select input parameter by number
    1. Mass Flow""")
    mode=input()
    if(mode=="1") or mode=="" or defaults==True:
        mdot=input("Input Desired Mass Flow in kg/s (default = .25)")
        if(mdot=="" or defaults==True):
            mdot=.25
        mdot=float(mdot)
        At=(mdot*Vt)/vt
        #print("At 1: " + (mdot/pt)*math.sqrt((Rspecific*Tt)/AverageMolecularWeight*k)) #wrong
        #print("At 2: " + str((mdot*Vt)/vt)) #derived from 3-24 on page 59 #correct
        #print("At 3: " + str((mdot*math.sqrt(T1)*((1/2)*(k-1)*(M2**2)+1)**((-1*k)/(2-2*k)-1/(2-2*k)))/(M2*p1*math.sqrt(k/Rspecific)))) #wrong
        #A2=At*(1/M2)*((1+((k-1)/2))/((k+1)/2))**((k+1)/(2*(k-1)))
        A2overAt=((1+M2**2*(k-1)/2)**((k+1)/(k-1)/2))*(((k+1)/2)**(-1*((k+1)/(k-1)/2)))/M2
        A2=A2overAt*At
        break
    else:
        print("Input not recognised. Please try again.")
    """
    elif mode=="2":
        r=float(input("Input Desired throat radius in meters"))
        At=pi*(r**2) #area of a circle from radius
        A2=At/((((k+1)/2)**(1/(k-1)))*((p2/p1)**(1/k))*math.sqrt(((k+1)/(k-1))*(1-((p2/p1)**((k-1)/k)))))
        mdot=At*p1*k*((2/(k+1))**((k+1)/(k-1)))/math.sqrt(k*Rspecific*T1)
        break"""



F = mdot * v2 + (p2 - p3) * A2

print("Force: "+str(round(F, 4))+"\n")

print("Mass Flow Rate: "+str(round(mdot, 4)))
print("Mach at exit: "+str(round(M2, 4))+"\n")

print("p1: "+str(round(p1, 4))) #given
print("T1: "+str(round(T1, 4))+"\n") #given

print("vt: "+str(round(vt, 6))) #HH p10
print("pt: "+str(round(pt, 6))) #HH p9 close nuf
print("At: "+str(round(At, 8))) #HH p9
print("Tt: "+str(round(Tt, 6))+"\n") #HH pg 9

print("v2: "+str(round(v2, 6))) #HH pg 10 close nuf
print("p2: "+str(round(p2, 4))) #given
print("A2: "+str(round(A2, 8))) #HH p10
print("T2: "+str(round(T2, 6))+"\n") #HH p9

print("p3: "+str(round(p3, 4))) #given

print("p2/p1: "+str(round(p2/p1, 6))) #HH
print("A2/At: "+str(round(A2/At, 8))) ##HH pg8
print("Tt/T2: "+str(round(Tt/T2, 6))) #fixed probably?

#print("Pc-injector/Pc-stagnation")

print("Specific Gas constant: " + str(round(Rspecific,6)))

genfileq=input("Generate a file? (yes/no)").lower()
if(genfileq=="yes"):
    print("look up desired Rao exit and throat angles. https://i.imgur.com/gq6j8UO.png")

    wb = openpyxl.load_workbook('partgen.xlsx') #name of the included file
    ws = wb.active #default sheet?

    ws['A1']="Parameter Name"
    ws['B1']="Value"
    ws['C1']="Units"

    ws['B8'] = str(float(input("Exit angle in deg:")))
    ws['c8'] = "deg"

    ws['B3'] = str(float(input("Throat angle in deg:")))
    ws['c3'] = "deg"

    ws['B4'] = str(str(A2)) #area exit
    ws['c4'] = "m^2"

    ws['B5'] = str(pi*(float(input("Combustion Chamber radius in INCHES:"))**2)) #area inlet
    ws['C5'] = "in^2"

    ws['B6'] = str(At)
    ws['C6'] = "m^2" #area throat

    ws['B7'] = str(float(input("Converging angle in degrees:")))
    ws['C7'] = "deg"

    #ws['B8'] = str(str(float(input("Rao Throat radius in INCHES:")))+" in")
    #ws['C8'] = "in"

    ws['B2'] = str(math.sqrt(At/3.14))
    ws['C2'] = "m"

    wb.save("partgenout.xlsx")
elif(genfileq.lower()=="no") or (genfileq=="" or defaults==True):
    print("Goodbye!")
else:
    print("Input not recognized")

#heat transfer and chamber
print("\nHeat Transfer:")
LOption=input("Attempt L* calculation?")
if(LOption.lower()=="yes"):
    print("Still in progress")#stub
elif(LOption.lower()=="no") or (LOption=="" or defaults==True):
    pass #LStar defined at top



Vc=LStar*At #chamber volume #rpe 287
print("Vc: "+str(round(Vc, 8)))
ts=Vc/(mdot*V1) #stay time #rpe 287
print("ts: "+str(round(ts, 4)))

Dc=input("Input chamber diameter: (default = 3\")")
if(Dc=="" or defaults==True):
    Dc=0.0508 #~2 in default chamber diameter
Dc=float(Dc)

A1=pi*((Dc/2)**2) #circles dawg

ConvergenceAngle=input("Input Convergence half angle in deg (default 30 deg): ")
if(ConvergenceAngle=="" or defaults==True):
    ConvergenceAngle=math.radians(30) #default convergence angle

Lc=Dc*math.tan(int(ConvergenceAngle)) #triangles dawg, pretty sure this is broke
L1=(Vc-A1*Lc*(1+math.sqrt(At/A1)+At/A1))/A1 #modified from RPE pg 285

print("L1, including conical: "+str(L1))
print("L1, not including conical:"+str(Vc/A1))

OFratio=input("input ox/fuel ratio")
if(OFratio=="" or defaults==True):
    OFratio=4
OFratio=float(OFratio)
mdot_fuel=mdot/(OFratio+1)
mdot_ox=mdot_fuel*OFratio

print("Ox mass flow rate: "+str(mdot_ox))
print("Fuel mass flow rate: "+str(mdot_fuel))

cdoxq=input("Input Cd for Ox")
if(cdoxq=="" or defaults==True):
    Cd_ox = 6.4
else:
    Cd_ox = float(cdoxq)

cdfuelq=input("Input Cd for fuel")
if(cdfuelq=="" or defaults==True):
    Cd_fuel = 6.4
else:
    Cd_fuel = float(cdfuelq)

oxdensityq=input("Input ox density")
if(oxdensityq=="" or defaults==True):
    ox_density=rho_liquid(initNoxTemp) #this assumes no multiphase flow
else:
    ox_density = float(oxdensityq)

fueldensityq=input("Input fuel density")
if(fueldensityq=="" or defaults==True):
    fuel_density=789
else:
    fuel_density = float(fueldensityq)

injectionPressure=2895798 #420 psi

OxInjectorArea=mdot_ox/(Cd_ox*math.sqrt(2*ox_density*(injectionPressure-p1)))
FuelInjectorArea=mdot_fuel/(Cd_fuel*math.sqrt(2*fuel_density*(injectionPressure-p1)))

print("Ox injector area: "+str(OxInjectorArea))
print("Fuel injector area: "+str(FuelInjectorArea))

oxportq=input("Input number of Ox ports")
if(oxportq=="" or defaults==True):
    OxPortNumber=2
else:
    OxPortNumber = int(oxportq)

fuelportq=input("Input number of Fuel ports")
if(fuelportq=="" or defaults==True):
    FuelPortNumber=1
else:
    FuelPortNumber = int(fuelportq)

print("\n"+"Ox injector diameter METERS"+str(2*math.sqrt((OxInjectorArea/OxPortNumber)/pi)))
print("Fuel injector diameter METERS"+str(2*math.sqrt((FuelInjectorArea/FuelPortNumber)/pi)))
print("Ox injector diameter INCHES"+str(Meters_to_Inches(2*math.sqrt((OxInjectorArea/OxPortNumber)/pi))))
print("Fuel injector diameter INCHES"+str(Meters_to_Inches(2*math.sqrt((FuelInjectorArea/FuelPortNumber)/pi)))+"\n")

ox_injection_velocity=Cd_ox*math.sqrt(2*(injectionPressure-p1)/ox_density) #rpe 282
fuel_injection_velocity=Cd_fuel*math.sqrt(2*(injectionPressure-p1)/fuel_density)
ox_angle = 8 #degrees
oxMomentum = mdot_ox*ox_injection_velocity*math.sin(math.radians(ox_angle)) #rpe pg282
fuel_angle = math.degrees(math.asin(oxMomentum/(mdot_fuel*fuel_injection_velocity)))
print("Ox injection angle from normal: "+str(ox_angle))
print("Fuel injection angle from normal: "+str(fuel_angle)+"\n")

DeltaPQuestion=input(">INPUT< Pressure drop or >CALC< pressure drop")
if(DeltaPQuestion.lower()=="input"):
    DeltaP_regen=input("Input Pressure Drop due to cooling passages")
    if(DeltaP_regen=="" or defaults==True):
        DeltaP_regen=p1*.1 #default of 10% of chamber pressure
elif(DeltaPQuestion=="calc" or DeltaPQuestion=="" or defaults==True):
    LenCoolantPipe=input("Input Length of coolant passage (default = 3\"): ")
    if(LenCoolantPipe=="" or defaults==True):
        LenCoolantPipe=0.0584 #~3in to m default passage length
    eqpipeq=input(">INPUT< or >CALC< equivilent pipe diameter: ")
    if("input"==eqpipeq) or (eqpipeq=="" or defaults==True):
        eqivD=input("Input equivilent pipe diameter (default = .00699 m): ")
        if(eqivD=="" or defaults==True):
            eqivD=0.00699625817
    elif("calc"==eqpipeq):
        #https://www.engineeringtoolbox.com/equivalent-diameter-d_205.html
        print("for a rectangular passage")
        a=input("input length of side one (default=.25\"): ")
        if(a=="" or defaults==True):
            a=0.0064 #.25 in to m
        a=float(a)
        b=input("input length of side two (default=.25\"): ")
        if (b == ""):
            b = 0.0064  # .25 in to m
        b=float(b)
        eqivD=(1.3*(a*b)**.625)/((a+b)**.25)
        print("Equivilent diameter" + str(eqivD))
    f_regen=input("Input friction loss coefficient (default = .03): ")
    if(f_regen=="" or defaults==True):
        f_regen=.03 #default friction loss coefficient #RPE pg 296
    f_regen=float(f_regen)
    v_regen = input("Input flow velocity (default = 2.109 m/s): ")
    if(v_regen=="" or defaults==True):
        v_regen=2.109
    v_regen=float(v_regen)
    if (v_regen == ""):
        v_regen = 1  # default flow velocity
    densityq=input(">INPUT< or >CALC< density")
    if(densityq=="input"):
        density=input("input your coolant density in kg/m^3")
    elif(densityq=="calc") or (densityq=="" or defaults==True):
        ox_temp=input("Input the temperature of OX storage (default 295)")
        if(ox_temp=="" or defaults==True):
            ox_temp=295
        ox_temp=float(ox_temp)
        density=(rho_liquid(ox_temp))
    DeltaP_regen=.5*(f_regen*v_regen**2)*(LenCoolantPipe/eqivD)*density
    print("Pressure drop: " + str(DeltaP_regen)+"\n")

print("now solving for gas film coeffcient")
ExhaustDensityq=input(">INPUT< or >CALC< exhaust density: ")
if(ExhaustDensityq=="calc") or (ExhaustDensityq=="" or defaults==True):
    print("ERROR: calc not currently working. Feeding default value")
    ExhaustDensity = 2.4794 #stub
else:
    ExhaustDensity=input("input exhaust density")
ExhaustDensity=float(ExhaustDensity)
print("Exhaust Density: " + str(ExhaustDensity))
Prandtl=input("Input Prandtl number (default=.72): ")
if(Prandtl=="" or defaults==True):
    Prandtl=.72 #https://onlinelibrary.wiley.com/doi/pdf/10.1002/0471722057.app2 pg 424
Prandtl=float(Prandtl)
GasConductivity=input("Input Gas Conductivity: ")
if(GasConductivity=="" or defaults==True):
    GasConductivity=0 #stub

AbsoluteGasViscosity=input("Input Absolute Gas Viscosity: ")
if(AbsoluteGasViscosity=="" or defaults==True):
    AbsoluteGasViscosity=(46.6*10**(-10))*((AverageMolecularWeight*2.20462)**.5)*((T1*9/5)**.6) #huzal and huang p86, adjusted for units
AbsoluteGasViscosity=float(AbsoluteGasViscosity)

chambervelocity=ox_injection_velocity

#GasFilmCoefficient = 0.023*(((ExhaustDensity*chambervelocity)**.8)/(Dc**.2))*(Prandtl**.4)*GasConductivity/(AbsoluteGasViscosity**.8)
#RoughGasFilmCoefficient=(ExhaustDensity*v1)**.8
rho_exhaust=p1/(Rspecific*T1)
# in progresssssssssssssssssssss #GasFilmCoefficient=(GasConductivity*.023*(((Dc*v*rho_exhaust)/AbsoluteGasViscosity)**.8)*(((AbsoluteGasViscosity*cp)/GasConductivity)**.4))/Dc


#GasFilmCoefficient=((.026/(Dc**.2))*(((AbsoluteGasViscosity**.2)*Cp)/(Prandtl**.6))*((p1/CStar)**.8)*((Dc*NozzleRadius)**.1))*(At/A1)*sigma #huzal pg 88

#term1=.0026/(Meters_to_Inches(Dc)**.2)
#term2=AbsoluteGasViscosity**2*Cp
#term3=
#term4=
#sigma=

#NozzleHeatTransfer=(term1*term2*(term3**.8)*(term4**.1))*sigma

#print("Gas Film Coefficient" + GasFilmCoefficient)
CoolantTemperature=input("Input Coolant Temperature: ")
if(CoolantTemperature=="" or defaults==True):
    CoolantTemperature=273 #stub
Cbar=input("Input Average Specific Heat of Coolant: ") #average specific heat
if(Cbar=="" or defaults==True):
    Cbar=C_liquid(CoolantTemperature)
CoolantVelocity=input("Input Coolant Velocity: ")
if(CoolantVelocity=="" or defaults==True):
    CoolantVelocity=1 #stub
Rho_coolant=input("Input Coolant Density: ")
if(Rho_coolant=="" or defaults==True):
    quality=0
    Rho_coolant=(1-quality)*rho_liquid(CoolantTemperature)+quality*rho_vapor(CoolantTemperature)
AbsoluteCoolantViscosity=input("Input Absolute Coolant Viscosity: ")
if(AbsoluteCoolantViscosity=="" or defaults==True):
    AbsoluteCoolantViscosity=DynViscosityLiquid(CoolantTemperature) #stub
CoolantConductivity=input("input Coolant Conductivity: ")
if(CoolantConductivity=="" or defaults==True):
    CoolantConductivity=ConductivityLiquid(CoolantTemperature) #stub
CoolantSurfaceArea=2*pi*(eqivD/2)*LenCoolantPipe
LiquidFilmCoefficient = 0.023*Cbar*(mdot_ox/CoolantSurfaceArea)*(((Dc*CoolantVelocity*Rho_coolant)/AbsoluteCoolantViscosity)**-2)*((AbsoluteCoolantViscosity*Cbar/CoolantConductivity)**(-2/3)) #rpe pg 318
print("Liquid Film Coefficient "+str(LiquidFilmCoefficient))

ChamberWallThickness=input("Input Chamber Wall Thickness ")
if(ChamberWallThickness=="" or defaults==True):
    ChamberWallThickness=0.0254
ChamberWallConductivity=input("Input Chamber Wall Conductivity: ")
if(ChamberWallConductivity=="" or defaults==True):
    ChamberWallConductivity=0 #stub
ConvectionHeatTransfer = (T1-CoolantTemperature)/(1/GasFilmCoefficient+ChamberWallThickness/ChamberWallConductivity+1/LiquidFilmCoefficient)
